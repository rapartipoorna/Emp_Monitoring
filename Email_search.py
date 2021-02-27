import imaplib
import pprint,pyzmail
from Email_login import imapobj,username

pprint.pprint(imapobj.list_folders())

imaplib._MAXLINE = 10000000

imapobj.select_folder('Inbox', readonly=True)
UIDs = imapobj.search(['SINCE', '24-Feb-2021', 'BEFORE', '25-Feb-2021'])
# imapobj.select_folder('[Gmail]/Sent Mail', readonly=True)
# UIDs_2 = imapobj.search(['SINCE', '24-Feb-2021', 'BEFORE', '25-Feb-2021'])
# UIDs.extend(UIDs_2)
# print(UIDs)
# print(UIDs_2)
# print(UIDs)
category = []
print(len(UIDs))
for i in range(len(UIDs)):
    label_dict = imapobj.get_gmail_labels(UIDs[i])
    print(label_dict)
    label = label_dict[UIDs[i]]
    if 'Starred' in str(label):
        category.append('Starred')
    elif 'Important' in str(label):
        category.append('Important')
    elif len(label) == 0:
        category.append('Inbox')
    else:
        category.append('Custom Label')
from_addresses = []
subjects = []
dates = []
days = []
months = []
years = []
times = []
sent_received = []
unsub_links = []
attachments=[]
print(len(UIDs))
for i in range(len(UIDs)):
    raw_message = imapobj.fetch(UIDs[i], ['BODY[]'])
    message = pyzmail.PyzMessage.factory(raw_message[UIDs[i]][b'BODY[]'])

    if message.get_address('from')[1] == username:
        sent_received.append('Sent')

    else:
        sent_received.append('Received')

    full_date = message.get_decoded_header('date')
    from_addresses.append(message.get_address('from'))
    subjects.append(message.get_subject(''))
    for part in message.walk():
        
        if bool(part.get_filename()):
            # attachments.append(part.get_filename())
            break
    attachments.append(part.get_filename())
        # print(str(i)+': '+str(part.get_filename()))
        # i+=1
    
    # print(str(i)+': '+str(message.get_filename()))
    # print(message.get_filename())
    unsub_link = message.get_decoded_header('List-Unsubscribe')
    if len(str(unsub_link)) > 0 and 'mailto' in unsub_link:
        unsub_link = unsub_link.split(',')
        unsub_links.append([unsub_link[idx] for idx, s in enumerate(unsub_link) if 'mailto' in s][0])
    else:
        unsub_links.append('No unsubscribe link found')

    day = full_date.split()[0].strip(',')
    date = full_date.split()[1]
    month = full_date.split()[2]
    year = full_date.split()[3]
    time = full_date.split()[4]

    days.append(day)
    dates.append(date)
    months.append(month)
    years.append(year)
    times.append(time)  

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "email_info"
ws.cell(1,1).value = "Date"
ws.cell(1,2).value = "Month"
ws.cell(1,3).value = "Year"
ws.cell(1,4).value = "Day"
ws.cell(1,5).value = "Time"
ws.cell(1,6).value = "From (Sender)"
ws.cell(1,7).value = "From (Email ID)"
ws.cell(1,8).value = "Subject"
ws.cell(1,9).value = "Sent/Received"
ws.cell(1,10).value = "Category"
ws.cell(1,11).value="Attachment"
for i in range(len(UIDs)):
    ws.cell(row=i+2, column=1).value = dates[i]
    ws.cell(row=i+2, column=2).value = months[i]
    ws.cell(row=i+2, column=3).value = years[i]
    ws.cell(row=i+2, column=4).value = days[i]
    ws.cell(row=i+2, column=5).value = times[i]
    ws.cell(row = i+2, column = 6).value = from_addresses[i][0]
    ws.cell(row = i+2, column = 7).value = from_addresses[i][1]
    ws.cell(row = i+2, column = 8).value = str(subjects[i])
    ws.cell(row=i + 2, column=9).value = sent_received[i]
    ws.cell(row=i+2, column=10).value = category[i]
    ws.cell(row=i+2, column=11).value = attachments[i]

wb.save('Email_Analytics.xlsx')          